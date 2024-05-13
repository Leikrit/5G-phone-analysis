from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import TimeoutException
from pyquery import PyQuery as pq
from urllib.parse import quote
import time
import openpyxl

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")  # 此处端口保持和命令行启动的端口一致
driver = Chrome(options=chrome_options, )  # executable_path='D:\Anaconda3\chrome-win64\chromedriver.exe'
wait = WebDriverWait(driver, 10)


wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)

ws.cell(row=1, column=1, value="index")
ws.cell(row=1, column=2, value="category")
ws.cell(row=1, column=3, value="picture_addr")


def index_page(ketword, max_page, count):
    search = ketword + '手机'
    url = "https://search.jd.com/Search?keyword=" + search
    driver.get(url)
    # btn = wait.until(EC.element_to_be_clickable(
    #     (By.XPATH, '//*[@id="J_selector"]/div[1]/div/div[2]/div[1]/ul/li[1]/a')))  # 参数按钮
    # btn.click()
    time.sleep(1.5)
    # 获取商品图片
    for i in range(max_page):
        driver.refresh()
        html = driver.page_source
        doc = pq(html)
        time.sleep(3.5)
        for y in range(28):
            js = 'window.scrollBy(0,200)'
            driver.execute_script(js)
            time.sleep(0.5)
        html = driver.page_source
        doc = pq(html)
        items = doc('.gl-i-wrap').items()  # 所有手机商品
        for item in items:
            print(item.find('.p-img')('img').attr('src'))
            product = {'category': keyword,
                       'picture_addr': item.find('.p-img')('img').attr('src')}
            if product['picture_addr'] != "None":
                product['picture_addr'] = "http:" + product['picture_addr']
            count += 1
            ws.cell(row=count, column=1, value=str(count - 1))
            ws.cell(row=count, column=2, value=str(product['category']))
            ws.cell(row=count, column=3, value=str(product['picture_addr']))
            print(product)
        if i != max_page-1:
            next_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="J_bottomPage"]/span[1]/a[9]')))  # 参数按钮
            next_btn.click()
            time.sleep(1.5)
    return count




def read_only():
    wb = openpyxl.load_workbook(filename="jd_picture.xlsx")
    ws = wb['Sheet1']
    lis = []
    dic = {"index": None, "category": None, "picture_addr": None}
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, values_only=True):  # , max_col=5
        dic["index"] = row[0]
        dic["category"] = row[1]
        dic["picture_addr"] = row[2]
        lis.append(dic)
    print(lis)
    return lis


if __name__ == '__main__':
    category = ['华为', '小米', 'apple', '荣耀', '三星', '红米', 'oppo', 'vivo', '一加', '魅族', 'iqoo', '真我']
    count = 1
    for i in category:
        keyword = i
        max_page = 3
        count = index_page(keyword, max_page, count)
    wb.save("jd_picture.xlsx")  # 保存手机商品的网址信息
