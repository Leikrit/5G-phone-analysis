from lxml import etree
import requests
import time
import openpyxl


def get_jd_item():
    wb = openpyxl.load_workbook(filename="jd.xlsx")
    ws = wb['Sheet1']
    ws.cell(row=1, column=5, value="specifications")
    count = 2
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=151, values_only=True):
        # print(row[0])
        ws.cell(row=count, column=5, value=jd_item(row[0]))
        count += 1
    wb.save("jd_item.xlsx")

def jd_item(id):
    url = "https://item.jd.com/" + str(id) + ".html"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
        "Cookie": "jsavif=1; shshshfpa=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; shshshfpx=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; __jdv=181111935|direct|-|none|-|1712498482836; __jdu=1712498482836958502465; areaId=19; thor=676741B8CA3FCE6E194F2D84F06C5CB3344018AEFA0977C08C33E05EF2AEFFC2E1A76BBB8EAEF54F167E0D5A091521C6C2647357A39A1425135A7B5D264B672461AA9B3FECC7E09DCA53EA4E31691ED9A536B6BBEC8FCD36A2255318F5481F71FAE3BDB5358DC3CB614F81F015A4A5493AFF45191AEBE70515C825B934B0DE5527DC7434634488AF35A89190257A5DB45A80FB5111C9A05CCD4F62CBD799AA35; flash=2_XCzzsYZ0FyMqOJoQdXzMtNQSvvqXN4q-SESB_0zOuXCIW6UYO6hBlJTCckeR68kqg0lk341UEqLcxbs8PCNHXTrPHmOUxb3pnoRtfkq1QpL*; pinId=BKmhqjK0Ot4FvWOdtn9TiA; pin=jd_lTmXtaLBvUsY; unick=jd_lTmXtaLBvUsY; ceshi3.com=000; _tp=7%2B9vkUgexJVrllJMczJ5lw%3D%3D; _pst=jd_lTmXtaLBvUsY; token=823960cc5f2c76f69f3dc03669daf4ec,3,951388; __tk=20128a4f95f573c29e9f0fc5f12745d2,3,951388; ipLoc-djd=19-1601-50258-129167; __jda=143920055.1712498482836958502465.1712498483.1712498483.1712498483.1; __jdc=143920055; user-key=d4e0b5c4-efd8-4e69-87e6-a0f85f57fe3a; cn=0; __jdb=143920055.11.1712498482836958502465|1.1712498483; 3AB9D23F7A4B3CSS=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMOXDZFCMQAAAAADWECUZTLHN6M6UX; _gia_d=1; shshshfpb=BApXeDsL6u-tAmC3SbfNMBEgpx2ei76hCBlAIg61t9xJ1MrE1XoO2; 3AB9D23F7A4B3C9B=2ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQ"
    }
    res = requests.get(url, headers=headers)
    res.encoding = 'utf-8'
    text = res.text
    selector = etree.HTML(text)
    list = selector.xpath('//*[@class="Ptable"]')
    for i in list:
        title = i.xpath('.//div[@class="Ptable-item"]/h3/text()')
        content1 = i.xpath('.//div[@class="Ptable-item"]/dl/dl/dt/text()')
        content2 = i.xpath('.//div[@class="Ptable-item"]/dl/dl/dd/text()')
        if len(content1) != len(content2):
            count = 0
            for i in content1:
                if i == '入网型号':
                    break
                count += 1
            content2[count:count + 3] = []
        print(id)
        print("title " + str(title))
        print("content1= " + str(content1))
        print("content2= " + str(content2))
        print("-----")
        content = ''
        for i in range(len(content1)):
            content += content1[i]
            content += ":"
            content += content2[i]
            content += ";"
        print(content)
        return content


# jd_item("100082427721")
get_jd_item()