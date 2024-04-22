# from lxml import etree
# import requests
# import time
# import openpyxl
#
# outwb = openpyxl.Workbook()
# outws = outwb.create_sheet(index=0)
#
# outws.cell(row=1, column=1, value="index")
# outws.cell(row=1, column=2, value="title")
# outws.cell(row=1, column=3, value="price")
# outws.cell(row=1, column=4, value="id")
#
# def jd_page():
#     page=1
#     s = 1
#     count = 1
#     for i in range(1,6):
#         print("page="+str(page)+",s="+str(s))
#         url = "https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想%5E&page="+str(page)+"&s="+str(s)+"&click=1"
#         page = page+1
#         s = s+30
#
#         headers = {
#             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
#             "Cookie": "jsavif=1; shshshfpa=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; shshshfpx=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; __jdv=181111935|direct|-|none|-|1712498482836; __jdu=1712498482836958502465; areaId=19; thor=676741B8CA3FCE6E194F2D84F06C5CB3344018AEFA0977C08C33E05EF2AEFFC2E1A76BBB8EAEF54F167E0D5A091521C6C2647357A39A1425135A7B5D264B672461AA9B3FECC7E09DCA53EA4E31691ED9A536B6BBEC8FCD36A2255318F5481F71FAE3BDB5358DC3CB614F81F015A4A5493AFF45191AEBE70515C825B934B0DE5527DC7434634488AF35A89190257A5DB45A80FB5111C9A05CCD4F62CBD799AA35; flash=2_XCzzsYZ0FyMqOJoQdXzMtNQSvvqXN4q-SESB_0zOuXCIW6UYO6hBlJTCckeR68kqg0lk341UEqLcxbs8PCNHXTrPHmOUxb3pnoRtfkq1QpL*; pinId=BKmhqjK0Ot4FvWOdtn9TiA; pin=jd_lTmXtaLBvUsY; unick=jd_lTmXtaLBvUsY; ceshi3.com=000; _tp=7%2B9vkUgexJVrllJMczJ5lw%3D%3D; _pst=jd_lTmXtaLBvUsY; token=823960cc5f2c76f69f3dc03669daf4ec,3,951388; __tk=20128a4f95f573c29e9f0fc5f12745d2,3,951388; ipLoc-djd=19-1601-50258-129167; __jda=143920055.1712498482836958502465.1712498483.1712498483.1712498483.1; __jdc=143920055; user-key=d4e0b5c4-efd8-4e69-87e6-a0f85f57fe3a; cn=0; __jdb=143920055.11.1712498482836958502465|1.1712498483; 3AB9D23F7A4B3CSS=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMOXDZFCMQAAAAADWECUZTLHN6M6UX; _gia_d=1; shshshfpb=BApXeDsL6u-tAmC3SbfNMBEgpx2ei76hCBlAIg61t9xJ1MrE1XoO2; 3AB9D23F7A4B3C9B=2ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQ"
#         }
#         res = requests.get(url, headers=headers)
#         # time.sleep(2)
#         res.encoding = 'utf-8'
#         text = res.text
#
#         selector = etree.HTML(text)
#         list = selector.xpath('//*[@id="J_goodsList"]/ul/li')
#
#         for i in list:
#             title = i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
#             price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
#             product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_", "")
#             print("title " + str(title))
#             print("price= " + str(price))
#             print("product_id= " + str(product_id))
#             print("-----")
#             count += 1
#             outws.cell(row=count, column=1, value=str(count - 1))
#             outws.cell(row=count, column=2, value=str(title))
#             outws.cell(row=count, column=3, value=str(price))
#             outws.cell(row=count, column=4, value=str(product_id))
#
#
# jd_page()
# outwb.save("xx.xls")  #保存


from lxml import etree
import requests
import time
import openpyxl

outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)

outws.cell(row=1, column=1, value="index")
outws.cell(row=1, column=2, value="id")
outws.cell(row=1, column=3, value="title")
outws.cell(row=1, column=4, value="price")

def jd_page():
    #  https://search.jd.com/Search?keyword=5g%E6%89%8B%E6%9C%BA&suggest=1.his.0.0&wq=5g%E6%89%8B%E6%9C%BA&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page=1&s=1&click=0
    #  https://search.jd.com/Search?keyword=5g%E6%89%8B%E6%9C%BA&suggest=1.his.0.0&wq=5g%E6%89%8B%E6%9C%BA&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page=3&s=56&click=0
    #  https://search.jd.com/Search?keyword=5g%E6%89%8B%E6%9C%BA&suggest=1.his.0.0&wq=5g%E6%89%8B%E6%9C%BA&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page=5&s=116&click=0
    #  https://search.jd.com/Search?keyword=5g%E6%89%8B%E6%9C%BA&suggest=1.his.0.0&wq=5g%E6%89%8B%E6%9C%BA&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page=7&s=176&click=0
    page=1
    s = 1
    count = 1
    for i in range(1,6):
        print("page="+str(page)+",s="+str(s))
        url = "https://search.jd.com/Search?keyword=5G手机&wq=5G手机&pvid=feb1e81188ca4a11964991bce3bbf228&isList=0&page="+str(page)+"&s="+str(s)+"&click=0"
        page = page+1
        s = s+30

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
            "Cookie": "jsavif=1; shshshfpa=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; shshshfpx=076e83c2-9fb4-a2b0-fae0-ae1019291998-1712498482; __jdv=181111935|direct|-|none|-|1712498482836; __jdu=1712498482836958502465; areaId=19; thor=676741B8CA3FCE6E194F2D84F06C5CB3344018AEFA0977C08C33E05EF2AEFFC2E1A76BBB8EAEF54F167E0D5A091521C6C2647357A39A1425135A7B5D264B672461AA9B3FECC7E09DCA53EA4E31691ED9A536B6BBEC8FCD36A2255318F5481F71FAE3BDB5358DC3CB614F81F015A4A5493AFF45191AEBE70515C825B934B0DE5527DC7434634488AF35A89190257A5DB45A80FB5111C9A05CCD4F62CBD799AA35; flash=2_XCzzsYZ0FyMqOJoQdXzMtNQSvvqXN4q-SESB_0zOuXCIW6UYO6hBlJTCckeR68kqg0lk341UEqLcxbs8PCNHXTrPHmOUxb3pnoRtfkq1QpL*; pinId=BKmhqjK0Ot4FvWOdtn9TiA; pin=jd_lTmXtaLBvUsY; unick=jd_lTmXtaLBvUsY; ceshi3.com=000; _tp=7%2B9vkUgexJVrllJMczJ5lw%3D%3D; _pst=jd_lTmXtaLBvUsY; token=823960cc5f2c76f69f3dc03669daf4ec,3,951388; __tk=20128a4f95f573c29e9f0fc5f12745d2,3,951388; ipLoc-djd=19-1601-50258-129167; __jda=143920055.1712498482836958502465.1712498483.1712498483.1712498483.1; __jdc=143920055; user-key=d4e0b5c4-efd8-4e69-87e6-a0f85f57fe3a; cn=0; __jdb=143920055.11.1712498482836958502465|1.1712498483; 3AB9D23F7A4B3CSS=jdd032ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQAAAAMOXDZFCMQAAAAADWECUZTLHN6M6UX; _gia_d=1; shshshfpb=BApXeDsL6u-tAmC3SbfNMBEgpx2ei76hCBlAIg61t9xJ1MrE1XoO2; 3AB9D23F7A4B3C9B=2ZJOWCZ7WY5TLDNLMCMU4UZPUPCMOTWAXIN3TSF6B3HTD5DRF55DEGHKITFZ6WJG6UTQJRXS2KAGJ4MUOIRCBZU3GQ"
        }
        res = requests.get(url, headers=headers)
        # time.sleep(2)
        res.encoding = 'utf-8'
        text = res.text

        selector = etree.HTML(text)
        list = selector.xpath('//*[@id="J_goodsList"]/ul/li')

        for i in list:
            title = i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
            price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
            product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_", "")
            print("title " + str(title))
            print("price= " + str(price))
            print("product_id= " + str(product_id))
            print("-----")
            count += 1
            outws.cell(row=count, column=1, value=str(count - 1))
            outws.cell(row=count, column=2, value=str(title))
            outws.cell(row=count, column=3, value=str(price))
            outws.cell(row=count, column=4, value=str(product_id))


jd_page()
outwb.save("5G_phone.xls")  #保存